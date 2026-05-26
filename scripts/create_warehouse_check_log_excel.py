from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

OUTPUT_FILE = "warehouse_order_check_log.xlsx"

ORDERS_HEADERS = [
    "Date",
    "Order Number",
    "Picker Name",
    "Checker Name",
    "Start Time",
    "Check Time",
    "Finish Time",
    "Duration Minutes",
    "Lines Count",
    "Order Size",
    "Heavy/Bulky Items",
    "Small Items Packed Correctly",
    "Error Found",
    "Error Type",
    "Fixed",
    "Status",
    "Comment",
]

ORDER_SIZES = ["S", "M", "L", "XL"]
YES_NO = ["Yes", "No"]
YES_NO_NA = ["Yes", "No", "N/A"]
ERROR_TYPES = [
    "None",
    "Wrong Quantity",
    "Wrong Item",
    "Missing Item",
    "Small Items Loose",
    "Bad Packing",
    "Damaged Item",
    "Other",
]
STATUSES = [
    "Checked",
    "Returned to Picker",
    "Ready for Packing",
    "Packed",
    "Shipped",
]
PICKERS_LIST_MAX_ROW = 500

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ERROR_ROW_FILL = PatternFill(fill_type="solid", fgColor="FDE9D9")
RETURNED_ROW_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")


def style_header_row(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22


def create_lists_sheet(wb: Workbook):
    ws = wb.create_sheet("Lists")

    ws["A1"] = "Order Sizes"
    ws["B1"] = "Yes/No"
    ws["C1"] = "Yes/No/N/A"
    ws["D1"] = "Error Types"
    ws["E1"] = "Statuses"

    style_header_row(ws)

    for idx, value in enumerate(ORDER_SIZES, start=2):
        ws.cell(row=idx, column=1, value=value)

    for idx, value in enumerate(YES_NO, start=2):
        ws.cell(row=idx, column=2, value=value)

    for idx, value in enumerate(YES_NO_NA, start=2):
        ws.cell(row=idx, column=3, value=value)

    for idx, value in enumerate(ERROR_TYPES, start=2):
        ws.cell(row=idx, column=4, value=value)

    for idx, value in enumerate(STATUSES, start=2):
        ws.cell(row=idx, column=5, value=value)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 20

    return ws


def create_pickers_db_sheet(wb: Workbook):
    ws = wb.create_sheet("Pickers_DB")
    ws["A1"] = "Picker Name"
    style_header_row(ws)
    ws.column_dimensions["A"].width = 24
    return ws


def create_orders_log_sheet(wb: Workbook):
    ws = wb.active
    ws.title = "Orders_Log"

    ws.append(ORDERS_HEADERS)
    ws.append([None] * len(ORDERS_HEADERS))

    style_header_row(ws)

    column_widths = {
        "A": 12,
        "B": 16,
        "C": 18,
        "D": 18,
        "E": 11,
        "F": 11,
        "G": 11,
        "H": 16,
        "I": 11,
        "J": 10,
        "K": 18,
        "L": 27,
        "M": 11,
        "N": 20,
        "O": 10,
        "P": 20,
        "Q": 28,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"

    table = Table(displayName="OrdersTable", ref=f"A1:Q{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    ws["A2"].number_format = "yyyy-mm-dd"
    ws["E2"].number_format = "hh:mm"
    ws["F2"].number_format = "hh:mm"
    ws["G2"].number_format = "hh:mm"

    ws["H2"] = "=IF(OR(E2=\"\",G2=\"\"),\"\",ROUND((G2-E2)*1440,0))"

    return ws


def add_orders_data_validation(orders_ws, lists_ws, pickers_ws) -> None:
    max_row = 1048576

    dv_order_size = DataValidation(
        type="list",
        formula1=f"={lists_ws.title}!$A$2:$A${1 + len(ORDER_SIZES)}",
        allow_blank=True,
    )
    dv_yes_no = DataValidation(
        type="list",
        formula1=f"={lists_ws.title}!$B$2:$B${1 + len(YES_NO)}",
        allow_blank=True,
    )
    dv_yes_no_na = DataValidation(
        type="list",
        formula1=f"={lists_ws.title}!$C$2:$C${1 + len(YES_NO_NA)}",
        allow_blank=True,
    )
    dv_error_types = DataValidation(
        type="list",
        formula1=f"={lists_ws.title}!$D$2:$D${1 + len(ERROR_TYPES)}",
        allow_blank=True,
    )
    dv_statuses = DataValidation(
        type="list",
        formula1=f"={lists_ws.title}!$E$2:$E${1 + len(STATUSES)}",
        allow_blank=True,
    )
    dv_pickers = DataValidation(
        type="list",
        formula1=f"={pickers_ws.title}!$A$2:$A${PICKERS_LIST_MAX_ROW}",
        allow_blank=True,
    )

    orders_ws.add_data_validation(dv_order_size)
    orders_ws.add_data_validation(dv_yes_no)
    orders_ws.add_data_validation(dv_yes_no_na)
    orders_ws.add_data_validation(dv_error_types)
    orders_ws.add_data_validation(dv_statuses)
    orders_ws.add_data_validation(dv_pickers)

    dv_pickers.add(f"C2:C{max_row}")
    dv_pickers.add(f"D2:D{max_row}")
    dv_order_size.add(f"J2:J{max_row}")
    dv_yes_no.add(f"K2:K{max_row}")
    dv_yes_no_na.add(f"L2:L{max_row}")
    dv_yes_no.add(f"M2:M{max_row}")
    dv_error_types.add(f"N2:N{max_row}")
    dv_yes_no_na.add(f"O2:O{max_row}")
    dv_statuses.add(f"P2:P{max_row}")


def add_orders_conditional_formatting(ws) -> None:
    target_range = "A2:Q1048576"

    error_found_rule = FormulaRule(
        formula=['$M2="Yes"'],
        fill=ERROR_ROW_FILL,
    )
    returned_rule = FormulaRule(
        formula=['$P2="Returned to Picker"'],
        fill=RETURNED_ROW_FILL,
    )

    ws.conditional_formatting.add(target_range, error_found_rule)
    ws.conditional_formatting.add(target_range, returned_rule)


def create_daily_summary_sheet(wb: Workbook):
    ws = wb.create_sheet("Daily_Summary")

    ws["A1"] = "Summary Date"
    ws["B1"] = ""
    ws["A1"].font = Font(bold=True)
    ws["B1"].number_format = "yyyy-mm-dd"

    labels = [
        "Total checked orders",
        "Orders with errors",
        "Error rate %",
        "Total order lines",
        "Average lines per order",
        "Average duration minutes",
        "S orders count",
        "M orders count",
        "L orders count",
        "XL orders count",
    ]

    for row_idx, label in enumerate(labels, start=3):
        ws.cell(row=row_idx, column=1, value=label)

    ws["B3"] = '=COUNTIFS(Orders_Log!$A:$A,$B$1,Orders_Log!$B:$B,"<>")'
    ws["B4"] = '=COUNTIFS(Orders_Log!$A:$A,$B$1,Orders_Log!$M:$M,"Yes")'
    ws["B5"] = "=IF(B3=0,0,B4/B3)"
    ws["B6"] = "=SUMIFS(Orders_Log!$I:$I,Orders_Log!$A:$A,$B$1)"
    ws["B7"] = "=IF(B3=0,0,B6/B3)"
    ws["B8"] = '=IFERROR(AVERAGEIFS(Orders_Log!$H:$H,Orders_Log!$A:$A,$B$1,Orders_Log!$H:$H,">0"),0)'
    ws["B9"] = '=COUNTIFS(Orders_Log!$A:$A,$B$1,Orders_Log!$J:$J,"S")'
    ws["B10"] = '=COUNTIFS(Orders_Log!$A:$A,$B$1,Orders_Log!$J:$J,"M")'
    ws["B11"] = '=COUNTIFS(Orders_Log!$A:$A,$B$1,Orders_Log!$J:$J,"L")'
    ws["B12"] = '=COUNTIFS(Orders_Log!$A:$A,$B$1,Orders_Log!$J:$J,"XL")'

    ws["B5"].number_format = "0.00%"

    style_header_row(ws)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18


def create_picker_summary_sheet(wb: Workbook):
    ws = wb.create_sheet("Picker_Summary")

    headers = [
        "Picker Name",
        "Total orders picked",
        "Orders with errors",
        "Error rate %",
        "Total lines picked",
        "Average lines per order",
        "Average duration minutes",
        "Returned orders count",
    ]
    ws.append(headers)
    style_header_row(ws)

    ws["A2"] = '=SORT(UNIQUE(FILTER(Orders_Log!$C$2:$C$1048576,Orders_Log!$C$2:$C$1048576<>"")))'

    for row in range(2, 1002):
        ws[f"B{row}"] = f'=IF($A{row}="","",COUNTIFS(Orders_Log!$C:$C,$A{row},Orders_Log!$B:$B,"<>"))'
        ws[f"C{row}"] = f'=IF($A{row}="","",COUNTIFS(Orders_Log!$C:$C,$A{row},Orders_Log!$M:$M,"Yes"))'
        ws[f"D{row}"] = f'=IF(B{row}=0,0,C{row}/B{row})'
        ws[f"E{row}"] = f'=IF($A{row}="","",SUMIFS(Orders_Log!$I:$I,Orders_Log!$C:$C,$A{row}))'
        ws[f"F{row}"] = f'=IF(B{row}=0,0,E{row}/B{row})'
        ws[f"G{row}"] = f'=IF($A{row}="","",IFERROR(AVERAGEIFS(Orders_Log!$H:$H,Orders_Log!$C:$C,$A{row},Orders_Log!$H:$H,">0"),0))'
        ws[f"H{row}"] = f'=IF($A{row}="","",COUNTIFS(Orders_Log!$C:$C,$A{row},Orders_Log!$P:$P,"Returned to Picker"))'
        ws[f"D{row}"].number_format = "0.00%"

    table = Table(displayName="PickerSummaryTable", ref="A1:H1001")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    widths = {
        "A": 22,
        "B": 18,
        "C": 18,
        "D": 12,
        "E": 16,
        "F": 20,
        "G": 22,
        "H": 20,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"


def create_error_summary_sheet(wb: Workbook):
    ws = wb.create_sheet("Error_Summary")

    ws.append(["Error type", "Count", "Percentage of all errors"])
    style_header_row(ws)

    error_types_for_summary = [et for et in ERROR_TYPES if et != "None"]

    for idx, error_type in enumerate(error_types_for_summary, start=2):
        ws.cell(row=idx, column=1, value=error_type)
        ws.cell(
            row=idx,
            column=2,
            value=f'=COUNTIFS(Orders_Log!$M:$M,"Yes",Orders_Log!$N:$N,A{idx})',
        )
        ws.cell(
            row=idx,
            column=3,
            value=f'=IF(SUM($B$2:$B${1 + len(error_types_for_summary)})=0,0,B{idx}/SUM($B$2:$B${1 + len(error_types_for_summary)}))',
        )
        ws.cell(row=idx, column=3).number_format = "0.00%"

    table = Table(displayName="ErrorSummaryTable", ref=f"A1:C{1 + len(error_types_for_summary)}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 24


def create_workbook(output_path: Path) -> None:
    wb = Workbook()

    orders_ws = create_orders_log_sheet(wb)
    lists_ws = create_lists_sheet(wb)
    pickers_ws = create_pickers_db_sheet(wb)
    add_orders_data_validation(orders_ws, lists_ws, pickers_ws)
    add_orders_conditional_formatting(orders_ws)

    create_daily_summary_sheet(wb)
    create_picker_summary_sheet(wb)
    create_error_summary_sheet(wb)

    wb.save(output_path)


def main() -> None:
    repo_root = Path(__file__).resolve().parent.parent
    output_path = repo_root / OUTPUT_FILE
    create_workbook(output_path)
    print(f"Created workbook: {output_path}")


if __name__ == "__main__":
    main()
